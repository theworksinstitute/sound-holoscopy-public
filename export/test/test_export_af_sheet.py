import pytest
from openpyxl.worksheet.worksheet import Worksheet

from export.config import template_path
from export.sheet_export.metasheet_af import export_af_sheet
from export.templates.FractionSheet import FractionSheet
from export.templates.AfSheet import AfSheet
from export.test.fixtures import export_config_stereo, export_config_surround_hacky, out_wb


@pytest.fixture
def template():
    return AfSheet(template_path)


@pytest.fixture
def frac_template():
    return FractionSheet(template_path)


def test_export_af_creates_sheet(export_config_stereo, template, frac_template, out_wb):
    export_af_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row,
                    out_wb,
                    lambda x: None)
    assert template.sheet_name in out_wb.sheetnames


# noinspection DuplicatedCode
def test_export_af_sheet_writes_header(export_config_stereo, template, frac_template, out_wb):
    export_af_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row,
                    out_wb,
                    lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]
    assert out_sheet.cell(1, template.xyz_position).value == 'xyz position'
    assert out_sheet.cell(1, template.f_range).value == 'f range'
    assert out_sheet.cell(1, template.avg_ad).value == 'average AD (-dB)'
    assert out_sheet.cell(1, template.avg_psi).value == 'average 𝜓 (degrees)'
    assert out_sheet.cell(1, template.max_ad).value == 'max AD (-dB)'
    assert out_sheet.cell(1, template.max_psi).value == 'max 𝜓 (degrees)'
    assert out_sheet.cell(1, template.min_ad).value == 'min AD (-dB)'
    assert out_sheet.cell(1, template.min_psi).value == 'min 𝜓 (degrees)'


def test_export_af_sheet_writes_data(export_config_stereo, template, frac_template, out_wb):
    export_af_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]

    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'
    assert out_sheet.cell(4, template.xyz_position).value == 'mics: _,2'
    assert out_sheet.cell(6, template.xyz_position).value == 'mics: 1,2'

    assert out_sheet.cell(2, template.f_range).value == '20-200'
    assert out_sheet.cell(3, template.f_range).value == '200-2000'
    assert out_sheet.cell(4, template.f_range).value == '20-200'

    assert out_sheet.cell(2, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H3,'fraction 2 of 3'!H3,'fraction 3 of 3'!H3)"
    assert out_sheet.cell(3, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H4,'fraction 2 of 3'!H4,'fraction 3 of 3'!H4)"
    assert out_sheet.cell(4, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H5,'fraction 2 of 3'!H5,'fraction 3 of 3'!H5)"
    assert out_sheet.cell(5, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H6,'fraction 2 of 3'!H6,'fraction 3 of 3'!H6)"
    assert out_sheet.cell(6, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H7,'fraction 2 of 3'!H7,'fraction 3 of 3'!H7)"
    assert out_sheet.cell(7, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H8,'fraction 2 of 3'!H8,'fraction 3 of 3'!H8)"

    assert out_sheet.cell(4, template.avg_psi).value == \
           "=AVERAGE('fraction 1 of 3'!U5,'fraction 2 of 3'!U5,'fraction 3 of 3'!U5)"
    assert out_sheet.cell(7, template.avg_psi).value == \
           "=AVERAGE('fraction 1 of 3'!U8,'fraction 2 of 3'!U8,'fraction 3 of 3'!U8)"


def test_export_af_sheet_surround_writes_data(export_config_surround_hacky, template, frac_template, out_wb):
    export_af_sheet(export_config_surround_hacky, frac_template.ad, frac_template.psi, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]

    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'
    assert out_sheet.cell(4, template.xyz_position).value == 'mics: _,2'
    assert out_sheet.cell(6, template.xyz_position).value == 'mics: 1,2'

    assert out_sheet.cell(2, template.f_range).value == '20-200'
    assert out_sheet.cell(3, template.f_range).value == '200-2000'
    assert out_sheet.cell(4, template.f_range).value == '20-200'

    assert out_sheet.cell(8, template.xyz_position).value == 'all mics'
    assert out_sheet.cell(8, template.f_range).value == '20-200'
    assert out_sheet.cell(8, template.avg_ad).value == '=AVERAGE(C2,C4,C6)'
    assert out_sheet.cell(8, template.max_ad).value == '=MAX(D2,D4,D6)'
    assert out_sheet.cell(8, template.avg_psi).value == '=AVERAGE(F2,F4,F6)'
    assert out_sheet.cell(8, template.max_psi).value == '=MAX(G2,G4,G6)'


# noinspection DuplicatedCode
def test_af_sheet_chart_exists(export_config_stereo, template, frac_template, out_wb):
    export_af_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]

    # Check that charts exist
    assert out_sheet._charts != []
    assert len(out_sheet._charts) == 2

    # Check that the charts have correct axis titles
    assert out_sheet._charts[0].y_axis.title.text.rich.p[0].text[0].value == 'AD (dB)'
    assert out_sheet._charts[0].x_axis.title.text.rich.p[0].text[0].value == 'f (Hz)'
    assert out_sheet._charts[1].y_axis.title.text.rich.p[0].text[0].value == '𝜓 (degrees)'
    assert out_sheet._charts[1].x_axis.title.text.rich.p[0].text[0].value == 'f (Hz)'

    # Check that there is correct amount of series
    assert len(out_sheet._charts[0].series) == 3 * 3  # 3 columns, 3 mic combinations
    assert len(out_sheet._charts[1].series) == 3 * 3  # 3 columns, 3 mic combinations
