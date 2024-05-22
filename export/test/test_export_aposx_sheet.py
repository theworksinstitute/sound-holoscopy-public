import pytest
from openpyxl.worksheet.worksheet import Worksheet

from export.config import template_path
from export.sheet_export.metasheet_aposx import export_aposx_sheet
from export.templates.AposxSheet import AposxSheet
from export.templates.AtSheet import AtSheet
from export.test.fixtures import export_config_mono, export_config_stereo, exported_mono, exported_stereo, out_wb


@pytest.fixture
def template():
    return AposxSheet(template_path)


@pytest.fixture
def at_template():
    return AtSheet(template_path)


def test_export_aposx_sheet_creates_sheet(export_config_stereo, template, at_template, out_wb):
    export_aposx_sheet(export_config_stereo, at_template.ad, at_template.psi_a, at_template.start_row, template,
                       out_wb, lambda x: None)
    assert 'Apos(x)' in out_wb.sheetnames


def test_export_aposx_sheet_writes_header(export_config_stereo, template, at_template, out_wb):
    export_aposx_sheet(export_config_stereo, at_template.ad, at_template.psi_a, at_template.start_row, template,
                       out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb['Apos(x)']
    assert out_sheet.cell(1, template.xyz_position).value == 'xyz position'
    assert out_sheet.cell(1, template.avg_ad).value == 'average AD (-dBA)'
    assert out_sheet.cell(1, template.max_ad).value == 'max AD (-dBA)'
    assert out_sheet.cell(1, template.min_ad).value == 'min AD (-dBA)'
    # ... and so on


def test_export_aposx_sheet_writes_data(export_config_stereo, template, at_template, out_wb):
    export_aposx_sheet(export_config_stereo, at_template.ad, at_template.psi_a, at_template.start_row,
                       template, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb['Apos(x)']
    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'

    assert out_sheet.cell(2, template.avg_ad).value == "=AVERAGE(At!E2:E4)"
    assert out_sheet.cell(2, template.max_ad).value == "=MAX(At!E2:E4)"
    assert out_sheet.cell(2, template.min_ad).value == "=MIN(At!E2:E4)"
    assert out_sheet.cell(2, template.avg_psi).value == "=AVERAGE(At!F2:F4)"
    assert out_sheet.cell(2, template.max_psi).value == "=MAX(At!F2:F4)"
    assert out_sheet.cell(2, template.min_psi).value == "=MIN(At!F2:F4)"

    assert out_sheet.cell(3, template.avg_ad).value == "=AVERAGE(At!E8:E10)"
    assert out_sheet.cell(3, template.max_ad).value == "=MAX(At!E8:E10)"
    assert out_sheet.cell(3, template.min_ad).value == "=MIN(At!E8:E10)"
    assert out_sheet.cell(3, template.avg_psi).value == "=AVERAGE(At!F8:F10)"
    assert out_sheet.cell(3, template.max_psi).value == "=MAX(At!F8:F10)"
    assert out_sheet.cell(3, template.min_psi).value == "=MIN(At!F8:F10)"


# noinspection DuplicatedCode
def test_aposx_sheet_chart_exists(export_config_stereo, template, at_template, out_wb):
    export_aposx_sheet(export_config_stereo, at_template.ad, at_template.psi_a, at_template.start_row, template,
                       out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb['Apos(x)']

    # Check that charts exist
    assert out_sheet._charts != []
    assert len(out_sheet._charts) == 2

    # Check that the charts have correct axis titles
    assert out_sheet._charts[0].y_axis.title.text.rich.p[0].text[0].value == 'AD (dB)'
    assert out_sheet._charts[0].x_axis.title.text.rich.p[0].text[0].value == 'pos (x)'
    assert out_sheet._charts[1].y_axis.title.text.rich.p[0].text[0].value == '𝜓 (degrees)'
    assert out_sheet._charts[1].x_axis.title.text.rich.p[0].text[0].value == 'pos (x)'

    # Check that there is correct amount of series
    assert len(out_sheet._charts[0].series) == 3  # avg, max and min AD
    assert len(out_sheet._charts[1].series) == 3  # avg, max and min PSI
