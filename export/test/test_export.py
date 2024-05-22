"""
This module tests integration of export module. It uses two sound files to generate output Excel file(s)
and then verifies that the cell values are within expected range.
"""
import os

from openpyxl.reader.excel import load_workbook

from export.config import sheets_dir
from export.test.fixtures import export_config_mono, export_config_stereo, exported_mono, exported_stereo, \
    exported_surround_hacky, export_config_surround_hacky, exported_surround, export_config_surround


# noinspection DuplicatedCode
def test_export_sheet_mono_fractions(exported_mono):
    """
    This test only verifies time fraction sheets for mono configuration, and does not touch meta-sheets.
    """

    test_wb = load_workbook(os.path.join(sheets_dir, 'test_mono.xlsx'))
    test_time_sheets = [sheet for sheet in test_wb.worksheets if "fraction" in sheet.title]
    exported_time_sheets = [sheet for sheet in exported_mono.worksheets if "fraction" in sheet.title]

    assert len(exported_time_sheets) == len(test_time_sheets)

    for i in range(len(test_time_sheets)):
        compare_worksheets(test_time_sheets[i], exported_time_sheets[i], 0.05)


def test_export_sheet_stereo_fractions(exported_stereo):
    """
    This test only verifies time fraction sheets for stereo configuration, and does not touch meta-sheets.
    """
    test_wb = load_workbook(os.path.join(sheets_dir, 'test_stereo.xlsx'))
    test_time_sheets = [sheet for sheet in test_wb.worksheets if "fraction" in sheet.title]
    exported_time_sheets = [sheet for sheet in exported_stereo.worksheets if "fraction" in sheet.title]

    assert len(exported_time_sheets) == len(test_time_sheets)

    for i in range(len(test_time_sheets)):
        compare_worksheets(test_time_sheets[i], exported_time_sheets[i], 0.05)


def test_export_sheet_mono_order(exported_mono):
    assert exported_mono.worksheets[0].title == 'Af'
    assert exported_mono.worksheets[1].title == 'At'
    assert exported_mono.worksheets[2].title == 'ft'


def test_export_sheet_stereo_order(exported_stereo):
    assert exported_stereo.worksheets[0].title == 'Af'
    assert exported_stereo.worksheets[1].title == 'At'
    assert exported_stereo.worksheets[2].title == 'Apos(x)'
    assert exported_stereo.worksheets[3].title == 'ft'
    assert exported_stereo.worksheets[4].title == 'fpos(x)'


def test_export_sheet_surround_order(exported_surround):
    assert exported_surround.worksheets[0].title == 'Af'
    assert exported_surround.worksheets[1].title == 'At'
    assert exported_surround.worksheets[2].title == 'Apos(x)'
    assert exported_surround.worksheets[3].title == 'Apos(y)'
    assert exported_surround.worksheets[4].title == 'ft'
    assert exported_surround.worksheets[5].title == 'fpos(x)'
    assert exported_surround.worksheets[6].title == 'fpos(y)'


def test_export_sheet_mono_ft(exported_mono):
    """
    This test only verifies "ft" sheet for mono configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_mono, 'ft', 'test_mono.xlsx', 0.05)


def test_export_sheet_stereo_ft(exported_stereo):
    """
    This test only verifies "ft" sheet for mono configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_stereo, 'ft', 'test_stereo.xlsx', 0.05)


def test_export_sheet_surround_hacky_ft(exported_surround_hacky):
    export_and_compare_worksheets(exported_surround_hacky, 'ft', 'test_surround.xlsx', 0.05)


def test_export_sheet_surround_ft(exported_surround):
    # TODO: Create a test for >2 mics, currently we only have hacky surround with 2 mics
    # export_and_compare_worksheets(exported_surround, 'ft', 'test_surround.xlsx', 0.05)
    pass


def test_export_sheet_mono_at(exported_mono):
    """
    This test only verifies "At" sheet for mono configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_mono, 'At', 'test_mono.xlsx', 0.05)


def test_export_sheet_stereo_at(exported_stereo):
    """
    This test only verifies "At" sheet for stereo configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_stereo, 'At', 'test_stereo.xlsx', 0.05)


def test_export_sheet_surround_hacky_at(exported_surround_hacky):
    """
    This test only verifies "At" sheet for surround configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_surround_hacky, 'At', 'test_surround.xlsx', 0.05)


def test_export_sheet_mono_aposx(exported_mono):
    """
    This test only verifies "Apos(x)" sheet for mono configuration, and does not touch other sheets.
    """
    assert 'Apos(x)' not in exported_mono.sheetnames


def test_export_sheet_stereo_aposx(exported_stereo):
    """
    This test only verifies "Apos(x)" sheet for stereo configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_stereo, 'Apos(x)', 'test_stereo.xlsx', 0.05)


def test_export_sheet_surround_hacky_aposx(exported_surround_hacky):
    """
    This test only verifies "Apos(x)" sheet for surround configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_surround_hacky, 'Apos(x)', 'test_surround.xlsx', 0.05)


def test_export_sheet_mono_fposx(exported_mono):
    """
    This test only verifies "fpos(x)" sheet for mono configuration, and does not touch other sheets.
    """
    assert 'fpos(x)' not in exported_mono.sheetnames


def test_export_sheet_stereo_fposx(exported_stereo):
    """
    This test only verifies "fpos(x)" sheet for stereo configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_stereo, 'fpos(x)', 'test_stereo.xlsx', 0.05)


def test_export_sheet_surround_hacky_fposx(exported_surround_hacky):
    """
    This test only verifies "fpos(x)" sheet for surround configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_surround_hacky, 'fpos(x)', 'test_surround.xlsx', 0.05)


def test_export_sheet_mono_af(exported_mono):
    """
    This test only verifies "Af" sheet for mono configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_mono, 'Af', 'test_mono.xlsx', 0.05)


def test_export_sheet_stereo_af(exported_stereo):
    """
    This test only verifies "Af" sheet for stereo configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_stereo, 'Af', 'test_stereo.xlsx', 0.05)


def test_export_sheet_surround_hacky_af(exported_surround_hacky):
    """
    This test only verifies "Af" sheet for surround configuration, and does not touch other sheets.
    """
    export_and_compare_worksheets(exported_surround_hacky, 'Af', 'test_surround.xlsx', 0.05)


# noinspection DuplicatedCode
def export_and_compare_worksheets(exported_wb, sheet_name, sheets_dir_name, tolerance=0.05):
    test_wb = load_workbook(os.path.join(sheets_dir, sheets_dir_name))
    test_sheet = test_wb[sheet_name]
    exported_sheet = exported_wb[sheet_name]
    compare_worksheets(test_sheet, exported_sheet, tolerance)


def compare_worksheets(ws1, ws2, tolerance=0.01):
    """
    A helper function that compares values between worksheets with specified tolerance.
    """
    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    for row_num in range(1, max_row + 1):
        for col_num in range(1, max_col + 1):
            cell1 = ws1.cell(row=row_num, column=col_num)
            cell2 = ws2.cell(row=row_num, column=col_num)
            if isinstance(cell1.value, (int, float)) and isinstance(cell2.value, (int, float)):
                assert abs(cell1.value - cell2.value) <= tolerance, \
                    f"Values at ({cell1.coordinate}) are not close enough."
            elif isinstance(cell1.value, str) and isinstance(cell2.value, str):
                assert cell1.value == cell2.value, f"Values at ({cell1.coordinate}) are not equal."
