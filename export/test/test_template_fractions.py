"""
This module tests integration of fraction template. It verifies that columns are found correctly, cells are copied
correctly, styles match, etc.
"""
import os

from openpyxl import load_workbook

import export.config
from export.templates.FractionSheet import FractionSheet
from export.config import template_path, sheets_dir


# This test exists because the template needed re-saving in Excel, otherwise it would not be read correctly
# See more about this issue:
# https://stackoverflow.com/questions/62800822/openpyxl-cannot-read-strict-open-xml-spreadsheet-format-userwarning-file-conta
# It had something to do with the area plots in the template. My solution was saving it directly from Google Docs
# and not re-saving with WPS Office or Excel.
def test_read_template():
    assert load_workbook(template_path)
    assert load_workbook(os.path.join(sheets_dir, 'test_mono.xlsx'))
    assert load_workbook(os.path.join(sheets_dir, 'test_stereo.xlsx'))


def test_template_fraction_columns():
    template = FractionSheet(template_path)
    assert template.s1_name == 1
    assert template.s2_name == 2
    assert template.range == 3
    assert template.f_hz_s1 == 4
    assert template.measured_s1 == 5
    assert template.f_hz_s2 == 6
    assert template.measured_s2 == 7
    assert template.f_hz_sd1 == 9
    assert template.measured_sd1 == 10
    assert template.f_hz_sd2 == 11
    assert template.measured_sd2 == 12


def test_copy_template_row():
    template = FractionSheet(template_path)
    template.copy_template_values(template.sheet, 10, 1)
    # template._wb.save(export.config.resources_dir + '/out.xlsx')
    assert template.sheet['H10'].value == '=E10-G10'


def test_copy_template_rows():
    template = FractionSheet(template_path)
    template.copy_template_values(template.sheet, 10, 7)
    # wb.save(resources_dir + '/out.xlsx')
    assert template.sheet['M16'].value == '=10^(E16/20)'
    assert template.sheet['M16'].border.bottom == template.sheet['M9'].border.bottom
    assert template.sheet['M16'].border.bottom != template.sheet['M15'].border.bottom


def test_copy_many_template_rows():
    template = FractionSheet(template_path)
    template.copy_template_values(template.sheet, 10, 100)
    # wb.save(resources_dir + '/out.xlsx')
    assert template.sheet['H109'].value == '=E109-G109'
    assert template.sheet['H10'].border.top == template.sheet['H3'].border.top
    assert template.sheet['H109'].border.bottom != template.sheet['H108'].border.bottom


def test_template_starting_row():
    template = FractionSheet(template_path)
    starting_row = template.start_row
    assert template.sheet[starting_row - 1][0].value == 's1'
    assert template.sheet[starting_row - 1][1].value == 's2'
    assert template.sheet[starting_row - 1][2].value == 'f range'
    assert template.sheet[starting_row][2].value == '50-52'
    assert template.sheet[starting_row + 1][2].value == '146-148'


def test_template_rows():
    template = FractionSheet(template_path)
    starting_row = template.start_row
    template_rows = template.template_rows
    assert template.sheet[starting_row - 1][0].value == 's1'
    assert template.sheet[starting_row][2].value == '50-52'
    assert template.sheet[starting_row + template_rows - 1][2].value == '2344-2346'
    assert template.sheet[starting_row + template_rows - 2][2].border.bottom is None
    assert template.sheet[starting_row + template_rows - 1][2].border.bottom.style is not None
    assert template.sheet[starting_row + template_rows][2].value is None
