import os
import shutil
from copy import deepcopy

import pytest
from openpyxl.reader.excel import load_workbook

from config import ExportConfig
from export.config import sounds_dir, template_path
from export.export import create_export

PRESERVE_EXPORT = True

mic1_c = os.path.join(sounds_dir, 'mic1_C.wav')
mic1_ref = os.path.join(sounds_dir, 'mic1_REF.wav')
mic2_c = os.path.join(sounds_dir, 'mic2_C.wav')
mic2_ref = os.path.join(sounds_dir, 'mic2_REF.wav')


@pytest.fixture(scope='session')
def export_config_mono():
    c_files = [mic1_c]
    ref_files = [mic1_ref]

    export_config = ExportConfig(c_files=c_files, ref_files=ref_files,
                                 c_name='C', ref_name='REF',
                                 time_fractions=3,
                                 export_audio=False,
                                 export_fft=False,
                                 frequency_regions=[(20, 200), (200, 2000)], json_load_path=None)

    yield export_config

    remove_files(export_config)


@pytest.fixture(scope='session')
def export_config_stereo():
    c_files = [mic1_c, mic2_c]
    ref_files = [mic1_ref, mic2_ref]

    export_config = ExportConfig(c_files=c_files, ref_files=ref_files,
                                 c_name='C', ref_name='REF',
                                 time_fractions=3,
                                 export_audio=False,
                                 export_fft=False,
                                 frequency_regions=[(20, 200), (200, 2000)], json_load_path=None)

    yield export_config

    remove_files(export_config)


@pytest.fixture(scope='session')
def export_config_surround():
    c_files = [mic1_c, mic2_c, mic1_c, mic2_c]
    ref_files = [mic1_ref, mic2_ref, mic1_ref, mic2_ref]

    export_config = ExportConfig(c_files=c_files, ref_files=ref_files,
                                 c_name='C', ref_name='REF',
                                 time_fractions=3,
                                 export_audio=False,
                                 export_fft=False,
                                 frequency_regions=[(20, 200), (200, 2000)], json_load_path=None)

    yield export_config

    remove_files(export_config)


@pytest.fixture(scope='session')
def export_config_surround_hacky(export_config_stereo):
    export_config_hacky = deepcopy(export_config_stereo)
    export_config_hacky._force_surround = True
    return export_config_hacky


@pytest.fixture(scope="module")
def exported_mono(export_config_mono):
    """
    This fixture sets up configuration for mono export and deletes that folder after the test is done.
    """
    regions = [(50, 52), (146, 148), (284, 286), (456, 458), (661, 663), (1279, 1281), (2344, 2346)]
    export_config_mono.frequency_regions = regions
    export_config_mono.time_fractions = 16
    create_export(export_config_mono, lambda x: None)
    exported_wb = load_workbook(export_config_mono.sheet_path())

    return exported_wb


@pytest.fixture(scope="module")
def exported_stereo(export_config_stereo):
    """
    This fixture sets up configuration for stereo export and deletes that folder after the test is done.
    """
    regions = [(50, 52), (146, 148), (284, 286), (456, 458), (661, 663), (1279, 1281), (2344, 2346)]
    export_config_stereo.frequency_regions = regions
    export_config_stereo.time_fractions = 16
    create_export(export_config_stereo, lambda x: None)
    exported_wb = load_workbook(export_config_stereo.sheet_path())

    return exported_wb


@pytest.fixture(scope="module")
def exported_surround_hacky(export_config_surround_hacky):
    regions = [(50, 52), (146, 148), (284, 286), (456, 458), (661, 663), (1279, 1281), (2344, 2346)]
    export_config_surround_hacky.frequency_regions = regions
    export_config_surround_hacky.time_fractions = 16
    create_export(export_config_surround_hacky, lambda x: None)
    exported_wb = load_workbook(export_config_surround_hacky.sheet_path())

    return exported_wb


@pytest.fixture(scope="module")
def exported_surround(export_config_surround):
    regions = [(50, 52), (146, 148), (284, 286), (456, 458), (661, 663), (1279, 1281), (2344, 2346)]
    export_config_surround.regions = regions
    export_config_surround.time_fractions = 16
    create_export(export_config_surround, lambda x: None)
    exported_wb = load_workbook(export_config_surround.sheet_path())
    return exported_wb


def remove_files(export_config):
    # These two lines are responsible for cleaning up
    # noinspection PyProtectedMember
    if not PRESERVE_EXPORT and os.path.exists(export_config._destination_folder):
        # noinspection PyProtectedMember
        shutil.rmtree(export_config._destination_folder)


@pytest.fixture
def out_wb():
    out_wb = load_workbook(template_path)
    # Remove all the sheets from out_wb
    for sheet in out_wb.worksheets:
        out_wb.remove(sheet)

    yield out_wb
