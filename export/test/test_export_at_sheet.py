import pytest
from openpyxl.worksheet.worksheet import Worksheet

from export.config import template_path
from export.sheet_export.metasheet_at import export_at_sheet
from export.templates.AtSheet import AtSheet
from export.templates.FractionSheet import FractionSheet
from export.test.fixtures import export_config_stereo, export_config_mono, export_config_surround_hacky, out_wb


@pytest.fixture
def template():
    return AtSheet(template_path)


@pytest.fixture
def frac_template():
    return FractionSheet(template_path)


def test_export_aposx_sheet_creates_sheet(export_config_stereo, template, frac_template, out_wb):
    export_at_sheet(export_config_stereo, frac_template.s1_p, frac_template.s2_p, frac_template.psi_1, template,
                    frac_template.start_row, out_wb, lambda x: None)
    assert template.sheet_name in out_wb.sheetnames


def test_export_at_sheet_stereo_writes_header(export_config_stereo, template, frac_template, out_wb):
    export_at_sheet(export_config_stereo, frac_template.s1_p, frac_template.s2_p, frac_template.psi_1, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]
    assert out_sheet.cell(1, template.t_fraction).value == 't fraction'
    assert out_sheet.cell(1, template.xyz_position).value == 'xyz position'
    assert out_sheet.cell(1, template.s1).value == 's1 (dBA)'
    assert out_sheet.cell(1, template.s2).value == 's2 (dBA)'
    # ... and so on


def test_export_at_sheet_surround_writes_header(export_config_surround_hacky, template, frac_template, out_wb):
    export_at_sheet(export_config_surround_hacky, frac_template.s1_p, frac_template.s2_p, frac_template.psi_1,
                    template, frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]
    # ...
    assert out_sheet.cell(1, template.avg_ad).value == 'average AD (-dBA)'
    assert out_sheet.cell(1, template.avg_psi).value == 'average 𝜓A (degrees)'
    assert out_sheet.cell(1, template.max_ad).value == 'max AD (-dBA)'
    assert out_sheet.cell(1, template.max_psi).value == 'max 𝜓A (degrees)'
    assert out_sheet.cell(1, template.min_ad).value == 'min AD (-dBA)'
    assert out_sheet.cell(1, template.min_psi).value == 'min 𝜓A (degrees)'


# noinspection DuplicatedCode
def test_export_at_sheet_writes_data(export_config_stereo, template, frac_template, out_wb):
    export_at_sheet(export_config_stereo, frac_template.s1_p, frac_template.s2_p, frac_template.psi_1, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]
    assert out_sheet.cell(2, template.t_fraction).value == 't1'
    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'
    assert out_sheet.cell(6, template.t_fraction).value == 't2'
    assert out_sheet.cell(6, template.xyz_position).value == 'mics: _,2'
    assert out_sheet.cell(10, template.t_fraction).value == 't3'
    assert out_sheet.cell(10, template.xyz_position).value == 'mics: 1,2'
    assert out_sheet.cell(2, template.s1).value == "=10*LOG(SUM('fraction 1 of 3'!Z3:Z4))"
    assert out_sheet.cell(2, template.s2).value == "=10*LOG(SUM('fraction 1 of 3'!AA3:AA4))"
    assert out_sheet.cell(2, template.ad).value == "=C2-D2"
    assert out_sheet.cell(2, template.psi_a).value == "=AVERAGE('fraction 1 of 3'!R3:R4)"


# noinspection DuplicatedCode
def test_export_at_sheet_surround_writes_data(export_config_surround_hacky, template, frac_template, out_wb):
    export_at_sheet(export_config_surround_hacky, frac_template.s1_p, frac_template.s2_p, frac_template.psi_1, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]
    assert out_sheet.cell(2, template.t_fraction).value == 't1'
    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'
    assert out_sheet.cell(6, template.t_fraction).value == 't2'
    assert out_sheet.cell(6, template.xyz_position).value == 'mics: _,2'
    assert out_sheet.cell(10, template.t_fraction).value == 't3'
    assert out_sheet.cell(10, template.xyz_position).value == 'mics: 1,2'

    assert out_sheet.cell(11, template.avg_ad).value == "=AVERAGE(E2,E5,E8)"
    assert out_sheet.cell(12, template.avg_psi).value == "=AVERAGE(F3,F6,F9)"
    assert out_sheet.cell(13, template.max_ad).value == "=MAX(E4,E7,E10)"
    assert out_sheet.cell(11, template.max_psi).value == "=MAX(F2,F5,F8)"
    assert out_sheet.cell(12, template.min_ad).value == "=MIN(E3,E6,E9)"
    assert out_sheet.cell(13, template.min_psi).value == "=MIN(F4,F7,F10)"


def test_at_sheet_chart_exists(export_config_stereo, template, frac_template, out_wb):
    export_at_sheet(export_config_stereo, frac_template.ad, frac_template.psi, frac_template.psi_1, template,
                    frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb[template.sheet_name]

    # Check that charts exist
    assert out_sheet._charts != []
    assert len(out_sheet._charts) == 2

    # Check that the charts have correct axis titles
    assert out_sheet._charts[0].y_axis.title.text.rich.p[0].text[0].value == 'AD (dB)'
    assert out_sheet._charts[0].x_axis.title.text.rich.p[0].text[0].value == 't (min)'
    assert out_sheet._charts[1].y_axis.title.text.rich.p[0].text[0].value == '𝜓 (degrees)'
    assert out_sheet._charts[1].x_axis.title.text.rich.p[0].text[0].value == 't (min)'

    # Check that there is correct amount of series
    assert len(out_sheet._charts[0].series) == 1 * 3  # 1 column, 3 mic positions
    assert len(out_sheet._charts[1].series) == 1 * 3  # 1 column, 3 mic positions
