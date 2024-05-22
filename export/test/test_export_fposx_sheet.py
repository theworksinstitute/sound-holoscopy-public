import pytest
from openpyxl.worksheet.worksheet import Worksheet

from export.config import template_path
from export.sheet_export.metasheet_fposx import export_fposx_sheet
from export.templates.FposxSheet import FposxSheet
from export.templates.FractionSheet import FractionSheet
from export.test.fixtures import export_config_stereo, out_wb


@pytest.fixture
def template():
    return FposxSheet(template_path)


@pytest.fixture
def frac_template():
    return FractionSheet(template_path)


def test_export_fposx_sheet_creates_sheet(export_config_stereo, template, frac_template, out_wb):
    export_fposx_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                       lambda x: None)
    assert 'fpos(x)' in out_wb.sheetnames


def test_export_fposx_sheet_writes_header(export_config_stereo, template, frac_template, out_wb):
    export_fposx_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                       lambda x: None)
    out_sheet: Worksheet = out_wb['fpos(x)']
    assert out_sheet.cell(1, template.xyz_position).value == 'xyz position'
    assert out_sheet.cell(1, template.f_range).value == 'f range'
    assert out_sheet.cell(1, template.avg_ad).value == 'average AD (-dB)'
    assert out_sheet.cell(1, template.avg_psi).value == 'average 𝜓 (degrees)'
    assert out_sheet.cell(1, template.max_ad).value == 'max AD (-dB)'
    assert out_sheet.cell(1, template.max_psi).value == 'max 𝜓 (degrees)'
    assert out_sheet.cell(1, template.min_ad).value == 'min AD (-dB)'
    assert out_sheet.cell(1, template.min_psi).value == 'min 𝜓 (degrees)'


def test_export_fposx_sheet_writes_data(export_config_stereo, template, frac_template, out_wb):
    export_fposx_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                       lambda x: None)
    out_sheet: Worksheet = out_wb['fpos(x)']

    assert out_sheet.cell(2, template.xyz_position).value == 'mics: 1,_'
    assert out_sheet.cell(3, template.xyz_position).value == 'mics: 1,2'
    assert out_sheet.cell(4, template.xyz_position).value == 'mics: _,2'

    assert out_sheet.cell(2, template.f_range).value == '20-200'
    assert out_sheet.cell(5, template.f_range).value == '200-2000'

    assert out_sheet.cell(2, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H3,'fraction 2 of 3'!H3,'fraction 3 of 3'!H3)"
    assert out_sheet.cell(3, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H7,'fraction 2 of 3'!H7,'fraction 3 of 3'!H7)"
    assert out_sheet.cell(4, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H5,'fraction 2 of 3'!H5,'fraction 3 of 3'!H5)"
    assert out_sheet.cell(5, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H4,'fraction 2 of 3'!H4,'fraction 3 of 3'!H4)"
    assert out_sheet.cell(6, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H8,'fraction 2 of 3'!H8,'fraction 3 of 3'!H8)"
    assert out_sheet.cell(7, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H6,'fraction 2 of 3'!H6,'fraction 3 of 3'!H6)"
    assert out_sheet.cell(7, template.avg_psi).value == \
           "=AVERAGE('fraction 1 of 3'!U6,'fraction 2 of 3'!U6,'fraction 3 of 3'!U6)"


# noinspection DuplicatedCode
def test_fposx_sheet_chart_exists(export_config_stereo, template, frac_template, out_wb):
    export_fposx_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template,
                       frac_template.start_row, out_wb, lambda x: None)
    out_sheet: Worksheet = out_wb['fpos(x)']

    # Check that charts exist
    assert out_sheet._charts != []
    assert len(out_sheet._charts) == 2

    # Check that the charts have correct axis titles
    assert out_sheet._charts[0].y_axis.title.text.rich.p[0].text[0].value == 'AD (dB)'
    assert out_sheet._charts[0].x_axis.title.text.rich.p[0].text[0].value == 'pos (x)'
    assert out_sheet._charts[1].y_axis.title.text.rich.p[0].text[0].value == '𝜓 (degrees)'
    assert out_sheet._charts[1].x_axis.title.text.rich.p[0].text[0].value == 'pos (x)'

    # Check that there is correct amount of series
    assert len(out_sheet._charts[0].series) == 3 * 2  # 3 columns, 2 ranges
    assert len(out_sheet._charts[1].series) == 3 * 2  # 3 columns, 2 ranges
