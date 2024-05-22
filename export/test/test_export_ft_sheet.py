import pytest
from openpyxl.worksheet.worksheet import Worksheet

from export.config import template_path
from export.sheet_export.metasheet_ft import export_ft_sheet
from export.templates.FractionSheet import FractionSheet
from export.templates.FtSheet import FtSheet
from export.test.fixtures import export_config_stereo, out_wb


@pytest.fixture
def template():
    return FtSheet(template_path, mono=False)


@pytest.fixture
def frac_template():
    return FractionSheet(template_path)


def test_export_ft_sheet_creates_sheet(export_config_stereo, template, frac_template, out_wb):
    export_ft_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                    lambda x: None)
    assert 'ft' in out_wb.sheetnames


def test_export_ft_sheet_writes_header(export_config_stereo, template, frac_template, out_wb):
    export_ft_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                    lambda x: None)
    out_sheet: Worksheet = out_wb['ft']
    assert out_sheet.cell(1, template.t_fraction).value == 't fraction'
    assert out_sheet.cell(1, template.f_range).value == 'f range'
    # ... and so on


def test_export_ft_sheet_writes_data(export_config_stereo, template, frac_template, out_wb):
    export_ft_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                    lambda x: None)
    out_sheet: Worksheet = out_wb['ft']
    assert out_sheet.cell(2, template.t_fraction).value == 't1'
    assert out_sheet.cell(2, template.f_range).value == '20-200'
    assert out_sheet.cell(3, template.t_fraction).value == 't2'
    assert out_sheet.cell(3, template.f_range).value == '20-200'
    assert out_sheet.cell(4, template.t_fraction).value == 't3'
    assert out_sheet.cell(4, template.f_range).value == '20-200'
    assert out_sheet.cell(5, template.t_fraction).value == 't1'
    assert out_sheet.cell(5, template.f_range).value == '200-2000'
    assert out_sheet.cell(6, template.t_fraction).value == 't2'
    assert out_sheet.cell(6, template.f_range).value == '200-2000'
    assert out_sheet.cell(7, template.t_fraction).value == 't3'
    assert out_sheet.cell(7, template.f_range).value == '200-2000'

    assert out_sheet.cell(2, template.avg_ad).value == \
           "=AVERAGE('fraction 1 of 3'!H3,'fraction 1 of 3'!H5,'fraction 1 of 3'!H7)"
    assert out_sheet.cell(3, template.max_ad).value == \
           "=MAX('fraction 2 of 3'!H3,'fraction 2 of 3'!H5,'fraction 2 of 3'!H7)"
    assert out_sheet.cell(4, template.min_ad).value == \
           "=MIN('fraction 3 of 3'!H3,'fraction 3 of 3'!H5,'fraction 3 of 3'!H7)"

    assert out_sheet.cell(2, template.min_psi).value == \
           "=MIN('fraction 1 of 3'!U3,'fraction 1 of 3'!U5,'fraction 1 of 3'!U7)"
    assert out_sheet.cell(3, template.max_psi).value == \
           "=MAX('fraction 2 of 3'!U3,'fraction 2 of 3'!U5,'fraction 2 of 3'!U7)"
    assert out_sheet.cell(4, template.avg_psi).value == \
           "=AVERAGE('fraction 3 of 3'!U3,'fraction 3 of 3'!U5,'fraction 3 of 3'!U7)"


def test_ft_sheet_chart_exists(export_config_stereo, template, frac_template, out_wb):
    export_ft_sheet(export_config_stereo, frac_template.ad, frac_template.psi, template, frac_template.start_row, out_wb,
                    lambda x: None)
    out_sheet: Worksheet = out_wb['ft']

    # Check that charts exist
    assert out_sheet._charts != []
    assert len(out_sheet._charts) == 2

    # Check that the charts have correct axis titles
    assert out_sheet._charts[0].y_axis.title.text.rich.p[0].text[0].value == 'AD (dB)'
    assert out_sheet._charts[0].x_axis.title.text.rich.p[0].text[0].value == 't (min)'
    assert out_sheet._charts[1].y_axis.title.text.rich.p[0].text[0].value == '𝜓 (degrees)'
    assert out_sheet._charts[1].x_axis.title.text.rich.p[0].text[0].value == 't (min)'

    # Check that there is correct amount of series
    assert len(out_sheet._charts[0].series) == 3 * 2  # 3 columns, 2 regions
    assert len(out_sheet._charts[1].series) == 3 * 2  # 3 columns, 2 regions
