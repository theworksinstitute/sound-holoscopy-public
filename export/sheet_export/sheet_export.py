from typing import Literal

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import ExportConfig
from export import config
from export.sheet_export.metasheet_af import export_af_sheet
from export.sheet_export.metasheet_aposx import export_aposx_sheet
from export.sheet_export.metasheet_at import export_at_sheet
from export.sheet_export.metasheet_fposx import export_fposx_sheet
from export.sheet_export.metasheet_ft import export_ft_sheet
from export.sheet_export.utils import save_workbook, get_mic_str, move_sheets_in_front
from export.templates.AfSheet import AfSheet
from export.templates.AposxSheet import AposxSheet
from export.templates.AtSheet import AtSheet
from export.templates.FposxSheet import FposxSheet
from export.templates.FractionSheet import FractionSheet
from export.templates.FtSheet import FtSheet
from signal_processing import fft
from signal_processing.signals import Signal


def create_export(s1_sums, s2_sums, sd1_sums, sd2_sums, export_config: ExportConfig, log=print):
    """
    Main function that does sheet exporting. It calls template functions to copy the template and
    then writes the data to the output sheet. It fills only some of the columns, the rest are filled
    with formulas. Custom log function can be passed to print the progress.
    """

    # Set up the template and output workbooks
    fraction_template = FractionSheet(config.template_path)
    ft_template = FtSheet(config.template_path, mono=export_config.is_mono())
    at_template = AtSheet(config.template_path)
    aposx_template = AposxSheet(config.template_path)
    fposx_template = FposxSheet(config.template_path)
    af_template = AfSheet(config.template_path)
    out_wb: Workbook = load_workbook(config.template_path)

    # Remove all the sheets from out_wb
    for sheet in out_wb.worksheets:
        out_wb.remove(sheet)

    log("Exporting time fractions...")
    export_time_fractions(export_config, s1_sums, s2_sums, sd1_sums, sd2_sums, fraction_template, out_wb, log)

    log("Exporting meta sheets...")
    # Export "Af" sheet
    export_af_sheet(export_config, fraction_template.ad, fraction_template.psi, af_template,
                    fraction_template.start_row, out_wb, log)

    # Export "At" sheet
    export_at_sheet(export_config, fraction_template.s1_p, fraction_template.s2_p, fraction_template.psi_1,
                    at_template, fraction_template.start_row, out_wb, log)

    # Export "Apos(x/y/z)" sheet
    if not export_config.is_mono():
        dim: Literal['x', 'y', 'z']
        for dim in export_config.get_signal_sets_spatial().keys():
            export_aposx_sheet(export_config, at_template.ad, at_template.psi_a, at_template.start_row,
                               aposx_template, out_wb, log, dim=dim)
            export_fposx_sheet(export_config, fraction_template.ad, fraction_template.psi, fposx_template,
                               fraction_template.start_row, out_wb, log, dim=dim)

    # Export "ft" sheet
    export_ft_sheet(export_config, fraction_template.ad, fraction_template.psi, ft_template,
                    fraction_template.start_row, out_wb, log)

    # Move the last worksheet to become the first
    move_sheets_in_front(out_wb, export_config.sheet_path(), 'Af', 'At', 'Apos', 'ft', 'fpos')


def export_time_fractions(export_config: ExportConfig, s1_sums, s2_sums, sd1_sums, sd2_sums, template, out_wb, log):
    """
    Exports time fraction sheets. It loops through all the time fractions and writes the data to the output sheet.
    """
    # Extract all the needed fields from the data object
    signal_sets = export_config.get_signal_sets()
    fractions = export_config.time_fractions
    regions = export_config.frequency_regions
    mic_count = export_config.get_mic_count()
    c_name = export_config.c_name
    ref_name = export_config.ref_name
    output_path = export_config.sheet_path()

    # Loop through all the fractions
    for t in range(fractions):
        # Create a new sheet for time fractions in the output workbook
        time_str = f'fraction {t + 1} of {fractions}'

        out_sheet: Worksheet = out_wb.create_sheet(time_str)
        template.copy_template_header(out_sheet)

        for i in range(len(signal_sets)):
            # Actually copy template values in the output sheet
            row_count = len(regions)
            row = template.start_row + row_count * i
            template.copy_template_values(out_sheet, row, row_count)

            # Write the microphone numbers and C/REF names in the sheet
            out_sheet.cell(row + 2, 1).value = get_mic_str(signal_sets[i], mic_count)
            out_sheet.cell(row, template.s1_name).value = c_name
            out_sheet.cell(row, template.s2_name).value = ref_name

        log(f'fraction {t + 1}/{fractions}: analyzing {c_name} (0%)')
        write_data_for_set(out_sheet, s1_sums, regions, template.start_row, template.f_hz_s1,
                           template.measured_s1, template.range, t, fractions)

        log(f'fraction {t + 1}/{fractions}: analyzing {ref_name} (25%)')
        write_data_for_set(out_sheet, s2_sums, regions, template.start_row, template.f_hz_s2,
                           template.measured_s2, template.range, t, fractions)

        log(f'fraction {t + 1}/{fractions}: analyzing diff of {c_name} and {ref_name} (50%)')
        write_data_for_set(out_sheet, sd1_sums, regions, template.start_row, template.f_hz_sd1,
                           template.measured_sd1, template.range, t, fractions)

        log(f'fraction {t + 1}/{fractions}: analyzing diff+90 of {c_name} and {ref_name} (75%)')
        write_data_for_set(out_sheet, sd2_sums, regions, template.start_row, template.f_hz_sd2,
                           template.measured_sd2, template.range, t, fractions)

        log(f'fraction {t + 1}/{fractions}: analyzing done (100%)')

        save_workbook(out_wb, output_path)


def write_data_for_set(sheet: Worksheet, signal_set: [Signal],
                       regions: [int, int], starting_row: int,
                       frequency_col: int, amplitude_col: int,
                       harmonic_col: int, fraction_index: int, total_fractions: int):
    """
    Mutates the passed worksheet by filling in information about fft regions.

    :param sheet: Sheet to mutate.
    :param signal_set: Combination of signals from different microphones.
    :param regions: Frequency regions to export the information for.
    :param starting_row: Starting row index in the output sheet.
    :param frequency_col: Column index that holds max frequency values.
    :param amplitude_col: Column index that holds max amplitude values.
    :param harmonic_col: Column index that holds information about frequency ranges.
    :param fraction_index: Time fraction index to select from the signal set.
    :param total_fractions: Total number of time fractions.
    """
    for i in range(len(signal_set)):
        # Get the signal for the time fraction from the current microphone combination
        signal: Signal = signal_set[i].get_interval_fraction(total_fractions, fraction_index)

        full_fft = fft.create_fft(signal)
        row_count = len(regions)
        signal_row = starting_row + row_count * i
        for j in range(len(regions)):
            start, end = regions[j]
            s, e = full_fft.get_frequency_region(start, end)
            fft_region = full_fft.get_region(s, e)
            frequency, amplitude = fft_region.get_max_amplitude()
            sheet.cell(signal_row + j, frequency_col).value = frequency
            sheet.cell(signal_row + j, amplitude_col).value = amplitude
            sheet.cell(signal_row + j, harmonic_col).value = f'{start}-{end}'
            del fft_region
        del full_fft
