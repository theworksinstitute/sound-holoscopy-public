from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export.colors import get_odd_colors
from export.sheet_export.utils import save_workbook, get_mic_str, write_ad_psi_data, \
    create_ad_psi_elems_for_af_and_fposx, export_ad_psi_charts
from export.templates.AfSheet import AfSheet


def export_af_sheet(export_config, ad, psi, template: AfSheet, fractions_start_row, out_wb, log):
    """
    Exports the "Af" sheet. It loops through all mic positions and frequency regions, creates
    formulas for that and writes the data to the output sheet.
    """
    # Extract all the needed fields from the data object
    fractions = export_config.time_fractions
    regions = export_config.frequency_regions
    mic_count = export_config.get_mic_count()
    output_path = export_config.sheet_path()
    signal_sets = export_config.get_signal_sets()
    spatial_mapping = export_config.get_spatial_mapping_to_signal_set()
    is_surround = export_config.is_surround()

    out_sheet: Worksheet = out_wb.create_sheet(template.sheet_name)

    # Copy the header from the template
    log(f'Exporting "Af" table... (5/5)')
    template.copy_template_header(out_sheet)

    # Loop through all regions and fill the table
    extra_table = 1 if is_surround else 0
    for i in range(len(regions)):
        s, e = regions[i]

        # Loop through microphone combinations
        for j in range(len(signal_sets) + extra_table):
            is_surround_and_last = is_surround and j == len(signal_sets) + extra_table - 1

            # Copy styles
            copy_to_start = template.start_row + j * len(regions)
            template.copy_template_values(out_sheet, copy_to_start, len(regions), styles_only=True)

            # We'll use this for xyz_position column and referencing fraction sheet
            row_i = template.start_row + j * len(regions) + i

            # Create a full row in output sheet.
            if not is_surround_and_last:
                out_sheet.cell(row_i, template.xyz_position, get_mic_str(signal_sets[j], mic_count))
            else:
                out_sheet.cell(row_i, template.xyz_position, 'all mics')
            out_sheet.cell(row_i, template.f_range, f'{s}-{e}')

            if not is_surround_and_last:
                ad_elements, psi_elements = \
                    create_ad_psi_elems_for_af_and_fposx(ad, fractions, fractions_start_row, j, psi, regions, i)
                write_ad_psi_data(ad_elements, psi_elements, out_sheet, row_i, template)
            else:
                def get_col_elems(col):
                    return ','.join([f'{get_column_letter(col)}{template.start_row + x * len(regions) + i}'
                                     for x in range(len(signal_sets))])

                out_sheet.cell(row_i, template.avg_ad, f"=AVERAGE({get_col_elems(template.avg_ad)})")
                out_sheet.cell(row_i, template.max_ad, f"=MAX({get_col_elems(template.max_ad)})")
                out_sheet.cell(row_i, template.min_ad, f"=MIN({get_col_elems(template.min_ad)})")
                out_sheet.cell(row_i, template.avg_psi, f"=AVERAGE({get_col_elems(template.avg_psi)})")
                out_sheet.cell(row_i, template.max_psi, f"=MAX({get_col_elems(template.max_psi)})")
                out_sheet.cell(row_i, template.min_psi, f"=MIN({get_col_elems(template.min_psi)})")

    save_workbook(out_wb, output_path)

    log('Exporting "Af" charts... (5/5)')
    export_af_chart(out_sheet, out_wb, output_path, spatial_mapping, signal_sets, regions, template, is_surround)

    log('Exported "Af" metasheet! (5/5)')


def export_af_chart(out_sheet, out_wb, output_path, spatial_mapping, signal_sets, regions, template, is_surround):
    ad_chart = template.copy_template_chart(out_sheet, 0, 'f (Hz)', 'AD (dB)', ylim=(-20, 10))
    psi_chart = template.copy_template_chart(out_sheet, 1, 'f (Hz)', '𝜓 (degrees)', ylim=(-180, 180))

    color_generator = get_odd_colors()
    if not is_surround:
        for i in spatial_mapping:  # [0, 2, 1] for 2 microphones
            # Each microphone combination has a different color
            current_color = next(color_generator)

            # Variables that help select the rows for the current region
            s = template.start_row + i * len(regions)
            e = s + len(regions) - 1

            export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, current_color)
    else:
        color = next(color_generator)
        s = template.start_row + len(signal_sets) * len(regions)
        e = s + len(regions) - 1
        export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, color)

    save_workbook(out_wb, output_path)

