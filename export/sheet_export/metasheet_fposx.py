from typing import Literal

from openpyxl.worksheet.worksheet import Worksheet

from config import ExportConfig
from export.colors import get_odd_colors
from export.sheet_export.utils import get_mic_str, save_workbook, write_ad_psi_data, \
    export_ad_psi_charts, create_ad_psi_elems_for_af_and_fposx
from export.templates.FposxSheet import FposxSheet


def export_fposx_sheet(export_config: ExportConfig, ad, psi, template: FposxSheet, fractions_start_row, out_wb,
                       log=print, dim: Literal['x', 'y', 'z'] = 'x'):
    """
    Exports the "fpos(x)" sheet. It loops through all mic positions and frequency regions, creates
    formulas for that and writes the data to the output sheet.
    """
    # Extract all the needed fields from the data object
    fractions = export_config.time_fractions
    regions = export_config.frequency_regions
    mic_count = export_config.get_mic_count()
    output_path = export_config.sheet_path()
    signal_sets = export_config.get_signal_sets()
    spatial_mapping = export_config.get_spatial_mapping_to_signal_set(dim=dim)

    out_sheet: Worksheet = out_wb.create_sheet(f'fpos({dim})')

    # Copy the header from the template
    log(f'Exporting "fpos({dim})" table... (4/5)')
    template.copy_template_header(out_sheet)

    # Loop through all regions and fill the table
    row_i = template.start_row
    for i in range(len(regions)):
        s, e = regions[i]

        # Copy styles
        copy_to_start = template.start_row + i * len(spatial_mapping)
        template.copy_template_values(out_sheet, copy_to_start, len(spatial_mapping), styles_only=True)

        # Loop through microphone combinations
        for j in spatial_mapping:
            # Create a full row in output sheet.
            out_sheet.cell(row_i, template.xyz_position, get_mic_str(signal_sets[j], mic_count))
            out_sheet.cell(row_i, template.f_range, f'{s}-{e}')

            ad_elements, psi_elements = \
                create_ad_psi_elems_for_af_and_fposx(ad, fractions, fractions_start_row, j, psi, regions, i)

            write_ad_psi_data(ad_elements, psi_elements, out_sheet, row_i, template)

            row_i += 1

    save_workbook(out_wb, output_path)

    log(f'Exporting "fpos({dim})" charts... (4/5)')
    export_fposx_chart(out_sheet, out_wb, output_path, len(spatial_mapping), regions, template, dim)

    log(f'Exported "fpos({dim})" metasheet! (4/5)')


def export_fposx_chart(out_sheet, out_wb, output_path, mic_combinations_count, regions, template, dim):
    ad_chart = template.copy_template_chart(out_sheet, 0, f'pos ({dim})', 'AD (dB)', ylim=(-20, 10))
    psi_chart = template.copy_template_chart(out_sheet, 1, f'pos ({dim})', '𝜓 (degrees)', ylim=(-180, 180))

    color_generator = get_odd_colors()
    for i in range(len(regions)):
        # Each region has a different color
        current_color = next(color_generator)

        # Variables that help select the rows for the current region
        s = template.start_row + i * mic_combinations_count
        e = s + mic_combinations_count - 1

        export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, current_color)

    save_workbook(out_wb, output_path)
