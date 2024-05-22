from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export.colors import get_odd_colors
from export.sheet_export.utils import add_chart_data, save_workbook, get_mic_str, write_ad_psi_data, \
    export_ad_psi_charts
from export.templates.AtSheet import AtSheet


def export_at_sheet(export_config, s1_p, s2_p, psi_1, template: AtSheet, fractions_start_row, out_wb, log):
    """
    Exports the "At" sheet. It loops through all the time fractions and microphone combinations, creates
    formulas for that and writes the data to the output sheet.
    """
    fractions = export_config.time_fractions
    signal_sets = export_config.get_signal_sets()
    spatial_mapping = export_config.get_spatial_mapping_to_signal_set()
    regions = export_config.frequency_regions
    output_path = export_config.sheet_path()
    mic_count = export_config.get_mic_count()
    is_surround = export_config.is_surround()

    out_sheet: Worksheet = out_wb.create_sheet(template.sheet_name)

    log('Exporting "At" table... (2/5)')
    if not is_surround:
        template.copy_template_header(out_sheet, stop_col=template.psi_a)
    else:
        template.copy_template_header(out_sheet)

    # Loop through all the rows and fill the table
    row_i = template.start_row
    extra_set = 1 if is_surround else 0
    for i in range(len(signal_sets) + extra_set):
        is_last_set_and_surround = (i == len(signal_sets) + extra_set - 1 and is_surround)

        # Copy template row styles, but only copy values from AD column
        # Do not copy values after psi_a column, because those are used in surround, which is addressed below
        s = template.start_row + i * fractions  # Start row for current signal set
        if not is_last_set_and_surround:
            template.copy_template_values(out_sheet, s, fractions, num_cols=template.psi_a, styles_only=True)
            template.copy_template_column(out_sheet, s, template.ad, num_rows=fractions)
        else:
            # If it is last set, we copy only specific columns
            template.copy_template_values(out_sheet, s, fractions, num_cols=template.min_psi, styles_only=True)

        # Loop through all the time fractions
        for t in range(fractions):
            # Create a full row in output sheet
            out_sheet.cell(row_i, template.t_fraction, f't{t + 1}')
            if not is_last_set_and_surround:
                out_sheet.cell(row_i, template.xyz_position, get_mic_str(signal_sets[i], mic_count))
            else:
                out_sheet.cell(row_i, template.xyz_position, 'all mics')

            row_start = fractions_start_row + i * len(regions)
            row_end = row_start + len(regions) - 1

            s1_p_ch = get_column_letter(s1_p)
            s2_p_ch = get_column_letter(s2_p)
            psi_a_ch = get_column_letter(psi_1)
            if not is_last_set_and_surround:
                s1_value = f"=10*LOG(SUM('fraction {t + 1} of {fractions}'!{s1_p_ch}{row_start}:{s1_p_ch}{row_end}))"
                s2_value = f"=10*LOG(SUM('fraction {t + 1} of {fractions}'!{s2_p_ch}{row_start}:{s2_p_ch}{row_end}))"
                psi_a_value = f"=AVERAGE('fraction {t + 1} of {fractions}'!{psi_a_ch}{row_start}:{psi_a_ch}{row_end})"
                out_sheet.cell(row_i, template.s1, s1_value)
                out_sheet.cell(row_i, template.s2, s2_value)
                out_sheet.cell(row_i, template.psi_a, psi_a_value)
            else:
                # Self-reference AD and PSI columns
                ad_ch = get_column_letter(template.ad)
                psi_ch = get_column_letter(template.psi_a)
                ad_col_letters = [f'{ad_ch}{template.start_row + j * fractions + t}'
                                  for j in range(len(signal_sets))]
                psi_col_letters = [f'{psi_ch}{template.start_row + j * fractions + t}'
                                   for j in range(len(signal_sets))]
                write_ad_psi_data(ad_col_letters, psi_col_letters, out_sheet, row_i, template)

            row_i += 1
    save_workbook(out_wb, output_path)
    log('Exporting "At" charts... (2/5)')
    export_at_chart(out_sheet, out_wb, output_path, spatial_mapping, len(signal_sets), fractions, template, is_surround)
    log('Exported "At" metasheet! (2/5)')


def export_at_chart(out_sheet, out_wb, output_path, spatial_mapping, signal_sets_count, fractions, template,
                    is_surround):
    ad_chart = template.copy_template_chart(out_sheet, 0, 't (min)', 'AD (dB)', ylim=(-10, 5))
    psi_chart = template.copy_template_chart(out_sheet, 1, 't (min)', '𝜓 (degrees)', ylim=(0, 180))

    color_generator = get_odd_colors()
    if not is_surround:  # Will work for mono and stereo
        for i in spatial_mapping:  # [0, 2, 1] for 2 microphones
            # Each microphone combination has a different color
            current_color = next(color_generator)

            # Variables that help select the rows for the current region
            s = template.start_row + i * fractions
            e = s + fractions - 1

            add_chart_data(out_sheet, ad_chart, template.ad, s, e, current_color, solidLine=True)
            add_chart_data(out_sheet, psi_chart, template.psi_a, s, e, current_color, solidLine=True)
    else:  # Will work for surround
        color = next(color_generator)
        s = template.start_row + signal_sets_count * fractions
        e = s + fractions - 1

        export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, color)

    save_workbook(out_wb, output_path)
