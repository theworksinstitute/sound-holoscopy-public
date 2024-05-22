import os

from openpyxl.chart import Reference
from openpyxl.descriptors import Typed
from openpyxl.descriptors.nested import NestedInteger
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.drawing.colors import ColorChoice
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.xml.constants import DRAWING_NS


def save_workbook(wb: Workbook, file_name: str):
    """
    Saves the workbook to the given file name. Creates the directory if it doesn't exist.
    """
    try:
        os.makedirs(f'{os.path.dirname(file_name)}')
    except FileExistsError:
        pass
    wb.save(file_name)


# Helper function for setting the chart data.
# noinspection PyPep8Naming
def add_chart_data(sheet, chart, column, row_start, row_end, color, solidLine=True):
    """
    Adds the chart data as series, and color for the given chart.

    :param sheet: Sheet where the chart is.
    :param chart: Chart to add the data to.
    :param column: Column index that holds the data in the sheet.
    :param row_start: Row index where the data starts in the sheet.
    :param row_end: Row index where the data ends in the sheet.
    :param color: Color to use for the series. Must be in the format of "FF0000" (red).
    :param solidLine: Whether to use solid line or dashed line.
    """
    ref = Reference(sheet, min_col=column, max_col=column, min_row=row_start, max_row=row_end)
    chart.add_data(ref, titles_from_data=False)
    chart.series[-1].graphicalProperties.line.solidFill = color
    if not solidLine:
        chart.series[-1].graphicalProperties.line.dashStyle = "lgDash"
    else:
        chart.series[-1].graphicalProperties.line.width = 75000
    color_alpha = MyColorChoice(srgbClr=RGBAColor(color, alpha=15))
    chart.series[-1].graphicalProperties.solidFill = color_alpha


# Helper function for getting the microphone string from the signal set.
def get_mic_str(set_mics: [int], mic_count: int):
    """
    Returns a string that represents the microphone numbers in the signal set.
    For example, if the set_mics is [1] and mic_count is 2, the returned string will be "mics: 1,_".
    :param set_mics: The microphone numbers in the signal set.
    :param mic_count: The total number of microphones.
    """
    mics = ['_'] * mic_count
    for i in set_mics:
        mics[i - 1] = str(i)
    return 'mics: ' + ','.join(mics)


# Helper function for moving some sheets to the beginning of the workbook.
def move_sheets_in_front(out_wb: Workbook, output_path, *sheet_names):
    """
    Moves the sheets with the given names (not checked for strict equality) to the beginning.
    """
    repeat = 0
    for sheet_name in sheet_names:
        for i, wb_sheet_name in enumerate(out_wb.sheetnames):
            if wb_sheet_name.startswith(sheet_name):
                repeat += 1
                out_wb.move_sheet(out_wb[wb_sheet_name], offset=-i - 1 + repeat)
    save_workbook(out_wb, output_path)


# Helper class for creating transparent colors.
# https://stackoverflow.com/questions/66642773/creating-complex-area-chart-using-openpyxl-transparency
class RGBAColor(Serialisable):
    """
    RGBA color representation that can be assigned to area chart fill in order to achieve transparency.
    """
    tagname = 'srgbClr'
    namespace = DRAWING_NS
    val = Typed(expected_type=str)
    alpha = NestedInteger(allow_none=True)
    __elements__ = ('alpha',)

    def __init__(self, val, alpha=None):
        """
        :param alpha: Must be between 0 and 100. 100 is the same as solid fill.
        """
        self.val = val
        self.alpha = alpha * 1000


# Helper class for creating transparent colors.
# https://stackoverflow.com/questions/66642773/creating-complex-area-chart-using-openpyxl-transparency
class MyColorChoice(ColorChoice):
    srgbClr = Typed(expected_type=RGBAColor, allow_none=True)


def write_ad_psi_data(ad_elements, psi_elements, out_sheet, row_index, template):
    out_sheet.cell(row_index, template.avg_ad, f"=AVERAGE({','.join(ad_elements)})")
    out_sheet.cell(row_index, template.max_ad, f"=MAX({','.join(ad_elements)})")
    out_sheet.cell(row_index, template.min_ad, f"=MIN({','.join(ad_elements)})")
    out_sheet.cell(row_index, template.avg_psi, f"=AVERAGE({','.join(psi_elements)})")
    out_sheet.cell(row_index, template.max_psi, f"=MAX({','.join(psi_elements)})")
    out_sheet.cell(row_index, template.min_psi, f"=MIN({','.join(psi_elements)})")


def export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, start_row, end_row, color, export_min_max=True):
    add_chart_data(out_sheet, ad_chart, template.avg_ad, start_row, end_row, color)
    add_chart_data(out_sheet, psi_chart, template.avg_psi, start_row, end_row, color)
    if export_min_max:
        add_chart_data(out_sheet, ad_chart, template.max_ad, start_row, end_row, color, solidLine=False)
        add_chart_data(out_sheet, ad_chart, template.min_ad, start_row, end_row, color, solidLine=False)
        add_chart_data(out_sheet, psi_chart, template.max_psi, start_row, end_row, color, solidLine=False)
        add_chart_data(out_sheet, psi_chart, template.min_psi, start_row, end_row, color, solidLine=False)


def create_ad_psi_elems_for_af_and_fposx(ad, fractions, fractions_start_row, region_index, psi, regions, signal_index):
    ad_ch = get_column_letter(ad)
    psi_ch = get_column_letter(psi)
    frac_row = fractions_start_row + region_index * len(regions) + signal_index
    ad_elements = [f"'fraction {t + 1} of {fractions}'!{ad_ch}{frac_row}" for t in range(fractions)]
    psi_elements = [f"'fraction {t + 1} of {fractions}'!{psi_ch}{frac_row}" for t in range(fractions)]
    return ad_elements, psi_elements
