from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export.colors import get_odd_colors
from export.sheet_export.utils import save_workbook, get_mic_str, export_ad_psi_charts
from export.templates.AposxSheet import AposxSheet


def export_aposx_sheet(export_config, ad, psi_a, at_start_row, template: AposxSheet, out_wb, log=print, dim='x'):
    fractions = export_config.time_fractions
    signal_sets = export_config.get_signal_sets()
    output_path = export_config.sheet_path()
    mic_count = export_config.get_mic_count()
    spatial_mapping = export_config.get_spatial_mapping_to_signal_set(dim=dim)

    out_sheet: Worksheet = out_wb.create_sheet(f'Apos({dim})')

    log(f'Exporting "Apos({dim})" table... (3/5)')

    # Create header and copy styles
    template.copy_template_header(out_sheet)
    template.copy_template_values(out_sheet, at_start_row, len(spatial_mapping), styles_only=True)

    # Loop through all the rows and fill the table
    row_i = template.start_row
    for i in spatial_mapping:
        # Create a full row in output sheet
        s = at_start_row + i * fractions
        e = s + fractions - 1

        ad_letter = get_column_letter(ad)
        psi_letter = get_column_letter(psi_a)

        out_sheet.cell(row_i, template.xyz_position, get_mic_str(signal_sets[i], mic_count))
        out_sheet.cell(row_i, template.avg_ad, f"=AVERAGE(At!{ad_letter}{s}:{ad_letter}{e})")
        out_sheet.cell(row_i, template.max_ad, f"=MAX(At!{ad_letter}{s}:{ad_letter}{e})")
        out_sheet.cell(row_i, template.min_ad, f"=MIN(At!{ad_letter}{s}:{ad_letter}{e})")
        out_sheet.cell(row_i, template.avg_psi, f"=AVERAGE(At!{psi_letter}{s}:{psi_letter}{e})")
        out_sheet.cell(row_i, template.max_psi, f"=MAX(At!{psi_letter}{s}:{psi_letter}{e})")
        out_sheet.cell(row_i, template.min_psi, f"=MIN(At!{psi_letter}{s}:{psi_letter}{e})")

        row_i += 1

    save_workbook(out_wb, output_path)

    log(f'Exporting "Apos({dim})" charts... (3/5)')
    export_aposx_chart(out_sheet, out_wb, output_path, len(spatial_mapping), template, dim)

    log(f'Exported "Apos({dim})" metasheet! (3/5)')


def export_aposx_chart(out_sheet, out_wb, output_path, mic_combination_count, template, dim):
    ad_chart = template.copy_template_chart(out_sheet, 0, f'pos ({dim})', 'AD (dB)', ylim=(-10, 5))
    psi_chart = template.copy_template_chart(out_sheet, 1, f'pos ({dim})', '𝜓 (degrees)', ylim=(0, 180))

    color_generator = get_odd_colors()
    current_color = next(color_generator)

    # Variables that help select the rows for the current region
    s = template.start_row
    e = s + mic_combination_count - 1

    export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, current_color)

    save_workbook(out_wb, output_path)