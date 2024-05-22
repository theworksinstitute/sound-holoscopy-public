from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import ExportConfig
from export.colors import get_odd_colors
from export.config import SheetNamesFt
from export.sheet_export.utils import save_workbook, write_ad_psi_data, \
    export_ad_psi_charts
from export.templates.FtSheet import FtSheet


def export_ft_sheet(export_config: ExportConfig, ad_column: int, psi_column: int, template: FtSheet,
                    fractions_start_row, out_wb, log):
    """
    Exports the "ft" sheet. It loops through all the time fractions and writes the data to the output sheet.
    """
    # Extract all the needed fields from the data object
    fractions = export_config.time_fractions
    regions = export_config.frequency_regions
    mic_count = export_config.get_mic_count()
    output_path = export_config.sheet_path()
    signal_sets = export_config.get_signal_sets()
    is_mono = export_config.is_mono()

    out_sheet: Worksheet = out_wb.create_sheet('ft')

    log(f'Exporting "ft" table... (1/5)')
    # Copy the header from the template or create it manually for mono
    if is_mono:
        # Manually set header row for mono
        header_row = template.start_row - 1
        template.copy_template_values(out_sheet, header_row, 1, num_cols=4, styles_only=True)
        out_sheet.cell(header_row, template.t_fraction, SheetNamesFt.t_fraction.value)
        out_sheet.cell(header_row, template.f_range, SheetNamesFt.f_range.value)
        out_sheet.cell(header_row, template.avg_ad, SheetNamesFt.avg_ad.value)
        out_sheet.cell(header_row, template.avg_psi, SheetNamesFt.avg_psi.value)
    else:
        template.copy_template_header(out_sheet)

    # Loop through all regions and fill the table
    row_i = template.start_row
    for i in range(len(regions)):
        s, e = regions[i]

        # Copy styles
        num_cols_to_copy = 5 if mic_count == 1 else 40
        copy_to_start = template.start_row + i * fractions
        template.copy_template_values(out_sheet, copy_to_start, fractions, num_cols=num_cols_to_copy, styles_only=True)

        # Loop through time fractions
        for t in range(fractions):
            # Create a full row in output sheet.
            out_sheet.cell(row_i, template.t_fraction, f't{t + 1}')
            out_sheet.cell(row_i, template.f_range, f'{s}-{e}')

            # Mono and stereo have different column counts.
            ad_ch = get_column_letter(ad_column)
            psi_ch = get_column_letter(psi_column)

            rows_to_pick_from = [j * len(regions) + fractions_start_row + i
                                 for j in range(len(signal_sets))]
            ad_elements = [f"'fraction {t + 1} of {fractions}'!{ad_ch}{row}" for row in rows_to_pick_from]
            psi_elements = [f"'fraction {t + 1} of {fractions}'!{psi_ch}{row}" for row in rows_to_pick_from]

            if mic_count > 1:
                write_ad_psi_data(ad_elements, psi_elements, out_sheet, row_i, template)
            else:
                out_sheet.cell(row_i, template.avg_ad, f"={','.join(ad_elements)}")
                out_sheet.cell(row_i, template.avg_psi, f"={','.join(psi_elements)}")

            row_i += 1

    save_workbook(out_wb, output_path)

    log('Exporting "ft" charts... (1/5)')
    export_ft_chart(fractions, out_sheet, out_wb, output_path, regions, template, is_mono)

    log('Exported "ft" metasheet! (1/5)')


def export_ft_chart(fractions, out_sheet, out_wb, output_path, regions, template, is_mono):
    ad_chart = template.copy_template_chart(out_sheet, 0, 't (min)', 'AD (dB)', ylim=(-20, 10))
    psi_chart = template.copy_template_chart(out_sheet, 1, 't (min)', '𝜓 (degrees)', ylim=(-180, 180))

    color_generator = get_odd_colors()
    for i in range(len(regions)):
        # Each region has a different color
        current_color = next(color_generator)

        # Variables that help select the rows for the current region
        s = template.start_row + i * fractions
        e = s + fractions - 1

        export_ad_psi_charts(out_sheet, ad_chart, psi_chart, template, s, e, current_color, export_min_max=not is_mono)

        save_workbook(out_wb, output_path)
