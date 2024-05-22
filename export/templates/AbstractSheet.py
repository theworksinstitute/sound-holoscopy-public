"""
This file contains a class that is used to find column indices in the template. It is inherited by other classes that
represent specific sheets in the template. This class must not be instantiated directly.
"""
from abc import ABC, abstractmethod
from copy import copy, deepcopy

import pandas as pd
from openpyxl.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class AbstractSheet(ABC):
    """
    Class that manages finding column indices in the template. It is inherited by other classes that
    represent specific sheets in the template. This class must not be instantiated directly.
    """

    @abstractmethod
    def __init__(self, template_path, sheet_name, column_names_enum, start_row, template_rows):
        """
        Searches the corresponding column values in the config and dynamically adds their index
         as an attribute to this class. This class will have all the attributes that column name Enums,
         such as SheetNamesFraction and SheetNamesFt have.

        :param template_path: Path to the template.
        :param sheet_name: The name of the sheet in the template, e.g. "fraction", "ft", "fpos(x)"
        :param column_names_enum: The enum that contains the column names for the specific sheet.
        :param start_row: The row where the values start in the template. values_starting_row - 1 is the header row.
        :param template_rows: The number of rows that template holds. First, middle and last row are used for copying.
        """
        xls = pd.ExcelFile(template_path)
        df = pd.read_excel(xls, sheet_name, header=None)

        # Find the column indices in the specific sheet of the template
        for enum in column_names_enum:
            self.__setattr__(enum.name, self._find_attribute_in_excel(df.values, enum.value))

        self._wb = load_workbook(template_path)
        self.sheet_name = sheet_name
        self.sheet: Worksheet = self._wb[sheet_name]
        self.start_row = start_row
        self.template_rows = template_rows

    def copy_template_header(self, out_sheet: Worksheet, start_col=1, stop_col=40):
        """
        Copies the header row from template to output sheet.
        :param out_sheet: Destination sheet.
        :param start_col: Starting column index to copy values from. Default is 1.
        :param stop_col: Last column index to copy values from. Default is 40.
        """
        # Copy header row from template to output sheet
        for column in range(start_col, stop_col + 1):
            header_row = self.start_row - 1
            self._copy_cell(self.sheet.cell(header_row, column), out_sheet.cell(header_row, column), True)

    def copy_template_values(self, out_sheet: Worksheet, copy_to: int, num_rows: int = 4, start_col=1, num_cols=40,
                             styles_only=False):
        """
        Copies values from template sheet to destination. First and last rows are copied separately
        since they contain formatting (bold lines under or above the cell). Uses formula translator that
        modifies a formula so that it can be translated from one cell to another.

        However, this approach might not be what we want. For example, if we have a formula that is =A1+B1,
        and we copy it to the next row, it will become =A2+B2, but we might want it to be =A2+B1.
        For this reason we have an optional parameter that copies only cell styles and lets us fill in cell values
        by ourselves.

        :param out_sheet: Destination sheet.
        :param copy_to: Starting row index in out_sheet.
        :param num_rows: Number of rows to copy. Default is 4.
        :param start_col: Starting column index to copy values from. Default is 1.
        :param num_cols: Maximum number of columns to copy. Default is 40.
        :param styles_only: Whether to only copy styles.

        """
        copy_from = self.start_row
        for column in range(start_col, start_col + num_cols):
            # Copy the top row once
            self._copy_cell(self.sheet.cell(copy_from, column), out_sheet.cell(copy_to, column),
                            styles_only=styles_only)
            # Copy second row multiple times
            for i in range(1, num_rows - 1):
                row_to = copy_to + i
                self._copy_cell(self.sheet.cell(copy_from + 1, column), out_sheet.cell(row_to, column),
                                styles_only=styles_only)
            # Copy bottom row once
            self._copy_cell(self.sheet.cell(copy_from + self.template_rows - 1, column),
                            out_sheet.cell(copy_to + num_rows - 1, column), styles_only=styles_only)

    def copy_template_column(self, out_sheet: Worksheet, copy_to: int, column_index: int, num_rows: int = 4,
                             styles_only=False):
        """
        Same as copy_template_values, but only copies one column.
        """
        self.copy_template_values(out_sheet, copy_to, num_rows, start_col=column_index, num_cols=1, styles_only=styles_only)

    # noinspection PyProtectedMember
    def copy_template_chart(self, out_sheet, chart_index, x_title, y_title, xlim=None, ylim=None):
        """
        Copies the chart from the template to the output sheet. Automatically clears the series, so that
        the user can add their own data. Assumes that the chart exists in the template sheet.

        :return: The newly added chart in the destination sheet, to which a user can add their own data.
        :param out_sheet: Destination sheet.
        :param chart_index: Index of the chart in the template sheet.
        :param x_title: Title of the x-axis.
        :param y_title: Title of the y-axis.
        :param xlim: Limits of the x-axis. Tuple of (min, max).
        :param ylim: Limits of the y-axis. Tuple of (min, max).
        """

        # Create chart copies from the template
        copied_chart = deepcopy(self.sheet._charts[chart_index])
        out_sheet._charts.append(copied_chart)
        chart = out_sheet._charts[-1]

        # Set title properties, because openpyxl cannot read them from the template (or I couldn't find them)
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title

        # Clear the series, so that the user can add their own data
        chart.series = []

        if xlim is not None:
            chart.x_axis.scaling.min = xlim[0]
            chart.x_axis.scaling.max = xlim[1]
        if ylim is not None:
            chart.y_axis.scaling.min = ylim[0]
            chart.y_axis.scaling.max = ylim[1]

        return chart

    @staticmethod
    def _copy_cell(cell1: Cell, cell2: Cell, copy_value=False, styles_only=False):
        """
        Copies content and styling, translates formulae to adjust to new cell coordinates. This is limited to
        the same general restrictions of formulae: A1 cell-references only and no support for defined names.
        """
        # noinspection PyProtectedMember
        cell2._style = copy(cell1._style)
        if styles_only:
            return
        if cell1.data_type == 'f':
            cell2.value = Translator(cell1.value, origin=cell1.coordinate).translate_formula(cell2.coordinate)
        if copy_value:
            cell2.value = cell1.value

    def _find_attribute_in_excel(self, values: [[any]], column_name: str) -> int:
        column_indexes = map(lambda x: self._find_attribute_in_column(x, column_name), values)
        column_indexes = list(column_indexes)
        return next(x for x in column_indexes if x >= 0)

    @staticmethod
    def _find_attribute_in_column(values: [any], column_name: str) -> int:
        try:
            return next(i + 1 for (i, x) in enumerate(values) if x == column_name)
        except StopIteration:
            return -1
